#!/usr/bin/env python3
"""V7 SKU truth loader — extracts factual SKU data from listing.xlsx.

Schema follows v7_spec/c0_sku_truth_v7.yaml.

Workflow:
  listing.xlsx → parse rows by attribute label → regex-extract numbers / materials / use_cases → sku_truth dict

Coverage:
  - identity (sku_id, product_name_ru, category, archetype)
  - dimensions (cm/mm regex from Заголовок + Преимущества + Описание)
  - material (primary, finish — keyword matched)
  - use_cases (extracted from Применение / Преимущество 5)
  - canvas (default 3:4 / 1200x1600 unless overridden)
  - product_grade_anchor (forbidden_upgrade_keywords by archetype)

Limitations (require manual_input_required marker):
  - packaging.has_real_reference_image (needs visual confirmation of comm imagery)
  - quantity.units_per_pack (often implicit; let user confirm)
  - exact steel grade if not explicit

Usage:
  python3 sku_truth_loader.py --listing listing.xlsx --out sku_truth.yaml
  python3 sku_truth_loader.py --listing listing.xlsx --comm-dir 沟通图片 --out sku_truth.yaml
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# === Constants from v7 spec ===
CANVAS_DEFAULT = {
    "aspect_ratio": "3:4",
    "size_px": "1200x1600",
    "full_bleed": True,
}

# Each entry: (archetype_id, [keywords]). Detection picks the LONGEST matching keyword across all
# archetypes (longest match wins), so "канцеляр" (len 8) beats "коробк" (len 6) when both appear.
# This avoids substring false matches without manual priority gymnastics.
ARCHETYPE_KEYWORDS = [
    ("grooming_tool", ["ножницы", "маникюр", "педикюр", "пинцет", "пилка для ногт", "кусачк"]),
    ("office_craft", ["канцелярский нож", "канцеляр", "скальпел", "макетн", "резак", "нож для"]),
    ("kitchen_prep", ["кухонн", "повар", "разделочн", "лопатк", "венчик", "терк", "скалк"]),
    ("home_storage", ["контейнер", "органайзер", "коробк", "полк", "корзин"]),
    ("cosmetics", ["крем", "лосьон", "сыворот", "масло для", "помад", "тушь"]),
    ("small_electronics", ["электр", "зарядк", "наушник", "массаж", "фен"]),
    ("fashion_accessory", ["сумк", "ремень", "кошелек", "очки"]),
]

MATERIAL_KEYWORDS = {
    "metal": [
        ("нержавеющая сталь", "stainless steel"),
        ("углеродистая сталь", "carbon steel"),
        ("алюминий", "aluminum"),
        ("латунь", "brass"),
        ("сталь", "steel"),
    ],
    "plastic": [
        ("ABS", "ABS plastic"),
        ("полипропилен", "PP"),
        ("силикон", "silicone"),
        ("пластик", "plastic"),
    ],
    "fabric": [("хлопок", "cotton"), ("полиэстер", "polyester"), ("ткан", "fabric")],
    "ceramic": [("керамик", "ceramic"), ("фарфор", "porcelain")],
    "glass": [("стекл", "glass")],
    "electronic": [("датчик", "sensor"), ("аккумулятор", "battery")],
    "paper_wood": [("дерев", "wood"), ("бамбук", "bamboo"), ("картон", "cardboard")],
}

FINISH_KEYWORDS = [
    ("полированн", "polished"),
    ("матов", "matte"),
    ("шлифован", "brushed"),
    ("шероховат", "textured"),
    ("гладк", "smooth"),
    ("зеркальн", "mirror"),
]

FORBIDDEN_UPGRADE_KEYWORDS = [
    "luxury", "premium-luxury", "heirloom", "artisan", "jewelry-grade",
    "mirror-finish", "polished-to-mirror", "high-end", "boutique",
]

# Regex for dimensions: matches "8,5 см", "130 мм", "2.5 cm", "9 mm", etc.
# Russian uses comma as decimal sep, English uses period.
DIM_RE = re.compile(
    r"(\d{1,4}(?:[.,]\d{1,3})?)\s*(см|мм|cm|mm|sm)\b",
    re.IGNORECASE,
)


def _load_xlsx_rows(path: Path) -> List[Tuple[str, str, str]]:
    """Returns [(attr_label_ru, content_ru, content_zh)] from listing tab."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise SystemExit("❌ Need openpyxl: pip3 install --user openpyxl")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "listing" not in wb.sheetnames:
        raise SystemExit(f"❌ no 'listing' tab in {path}, sheets: {wb.sheetnames}")
    ws = wb["listing"]
    rows: List[Tuple[str, str, str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c) if c is not None else "" for c in row]
        if not any(c.strip() for c in cells):
            continue
        attr = cells[0].strip() if len(cells) > 0 else ""
        ru = cells[1].strip() if len(cells) > 1 else ""
        zh = cells[2].strip() if len(cells) > 2 else ""
        rows.append((attr, ru, zh))
    return rows


def _extract_dimensions(text_corpus: str) -> List[Dict[str, Any]]:
    """Find all dim hits (value, unit_normalized_to_cm)."""
    out: List[Dict[str, Any]] = []
    seen: set = set()
    for m in DIM_RE.finditer(text_corpus):
        raw_val = m.group(1).replace(",", ".")
        unit = m.group(2).lower()
        try:
            val = float(raw_val)
        except ValueError:
            continue
        # Normalize to cm per Q1 rule (≥1cm uses cm, <1cm uses mm)
        if unit in {"мм", "mm"}:
            cm = val / 10
            display = f"{val:g} мм" if val < 10 else f"{val/10:g} см"
        elif unit in {"см", "cm", "sm"}:
            cm = val
            display = f"{str(val).replace('.',',')} см"
        else:
            continue
        key = (round(cm, 3), unit)
        if key in seen:
            continue
        seen.add(key)
        out.append({
            "value_cm": round(cm, 3),
            "display_ru": display,
            "raw_token": m.group(0),
        })
    return out


def _detect_material(text: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Returns (category, ru_label, en_label)."""
    text_l = text.lower()
    for cat, options in MATERIAL_KEYWORDS.items():
        for ru_kw, en in options:
            if ru_kw.lower() in text_l:
                return cat, ru_kw, en
    return None, None, None


def _detect_finish(text: str) -> Optional[str]:
    text_l = text.lower()
    for ru_kw, en in FINISH_KEYWORDS:
        if ru_kw.lower() in text_l:
            return en
    return None


def _detect_archetype(text: str) -> Optional[str]:
    """Pick the archetype whose longest matching keyword wins (avoids substring traps)."""
    text_l = text.lower()
    best: Optional[Tuple[int, str]] = None  # (kw_len, archetype)
    for arch, keywords in ARCHETYPE_KEYWORDS:
        for kw in keywords:
            if kw.lower() in text_l and (best is None or len(kw) > best[0]):
                best = (len(kw), arch)
    return best[1] if best else None


def _extract_use_cases(rows: List[Tuple[str, str, str]]) -> List[Dict[str, str]]:
    """Pull use_cases from Применение / Преимущество 5 / Описание."""
    use_cases: List[Dict[str, str]] = []
    for attr, ru, _zh in rows:
        if not ru:
            continue
        # Преимущество 5 typically lists applications
        if "Применение" in attr or "Преимущество 5" in attr or attr == "Применение":
            # split on common separators (semicolons, commas at sentence boundaries)
            for chunk in re.split(r"[;。]|\.\s+", ru):
                chunk = chunk.strip().rstrip(".")
                if 5 <= len(chunk) <= 120:
                    use_cases.append({"case_ru": chunk, "source": f"listing_xlsx::{attr}"})
                    if len(use_cases) >= 6:
                        break
        if len(use_cases) >= 6:
            break
    return use_cases


def _row_lookup(rows: List[Tuple[str, str, str]], attr_substr: str) -> Optional[str]:
    for attr, ru, _zh in rows:
        if attr_substr.lower() in attr.lower() and ru:
            return ru
    return None


def build_sku_truth(
    listing_xlsx: Path,
    sku_id: Optional[str] = None,
    comm_dir: Optional[Path] = None,
    canvas_override: Optional[Dict[str, Any]] = None,
    comm_digest_max_images: int = 12,
    comm_digest_parallel: int = 3,
    skip_comm_digest: bool = False,
) -> Dict[str, Any]:
    rows = _load_xlsx_rows(listing_xlsx)
    title_ru = _row_lookup(rows, "Заголовок") or ""
    desc_ru = _row_lookup(rows, "Описание") or ""
    benefits_ru = " ".join(
        ru for attr, ru, _ in rows if "Преимущество" in attr and ru
    )
    full_corpus = " ".join([title_ru, desc_ru, benefits_ru])

    # Dimensions
    dim_hits = _extract_dimensions(full_corpus)
    dimensions: Dict[str, Any] = {"raw_hits": dim_hits, "source": "listing_xlsx"}

    # Material
    mat_cat, mat_ru, mat_en = _detect_material(full_corpus)
    material: Dict[str, Any] = {
        "category": mat_cat,
        "primary_ru": mat_ru,
        "primary_en": mat_en,
        "finish": _detect_finish(full_corpus),
        "source": "listing_xlsx" if mat_ru else "missing",
    }

    # Archetype
    archetype = _detect_archetype(full_corpus) or "generic"

    # Use cases
    use_cases = _extract_use_cases(rows)

    # Identity
    identity = {
        "sku_id": sku_id or listing_xlsx.parent.name,
        "product_name_ru": title_ru.split(":")[0].strip() if title_ru else "",
        "title_ru_full": title_ru,
        "description_ru": desc_ru,
        "category": "<inferred from archetype>",
        "archetype": archetype,
    }

    # Packaging — must be manually marked (we cannot detect real packaging from listing)
    has_real_pkg = False
    pkg_refs: List[str] = []
    if comm_dir and comm_dir.exists():
        for p in comm_dir.glob("*"):
            name = p.name.lower()
            if "упаков" in name or "packag" in name or "коробк" in name or "package" in name:
                has_real_pkg = True
                pkg_refs.append(str(p))
    packaging = {
        "has_real_reference_image": has_real_pkg,
        "reference_files": pkg_refs,
        "source": "comm_dir_scan" if comm_dir else "manual_input_required",
        "note": "If false, do NOT generate packaging/unboxing slot. Replace with scene-type slot.",
    }

    # Quantity (defaults; user should confirm)
    quantity = {
        "units_per_pack": 1,
        "pieces_per_unit": 1,
        "source": "manual_input_required",
        "note": "Default 1×1; verify against listing if multi-pack.",
    }

    # Grade anchor
    grade_anchor = {
        "market_segment": "mid-range",
        "finish_quality": "标准消费级",
        "forbidden_upgrade_keywords": FORBIDDEN_UPGRADE_KEYWORDS,
        "source": "default_template",
    }

    # Canvas
    canvas = dict(CANVAS_DEFAULT)
    if canvas_override:
        canvas.update(canvas_override)
    canvas["customer_override"] = canvas_override or None

    # Comm imagery digest (gpt-5.5 vision auto-extraction). Cached by content hash.
    comm_digest: Dict[str, Any] = {}
    if comm_dir and comm_dir.exists() and not skip_comm_digest:
        try:
            import comm_imagery_digest  # type: ignore
            comm_digest = comm_imagery_digest.scan_comm_dir(
                comm_dir,
                max_images=comm_digest_max_images,
                parallel=comm_digest_parallel,
            )
            # Promote selling_point_phrases that look like dimensions back into dimensions.raw_hits
            for sp in comm_digest.get("selling_point_phrases", []) or []:
                more = _extract_dimensions(sp)
                for h in more:
                    if not any(abs(h["value_cm"] - existing["value_cm"]) < 0.01 for existing in dim_hits):
                        h["from_comm_digest"] = True
                        dim_hits.append(h)
            dimensions["source"] = "listing_xlsx + comm_digest"
        except ImportError:
            comm_digest = {"error": "comm_imagery_digest module not importable"}
        except Exception as e:
            comm_digest = {"error": f"comm digest scan failed: {e}"[:200]}

    return {
        "schema_version": "v7.1",
        "canvas": canvas,
        "identity": identity,
        "dimensions": dimensions,
        "quantity": quantity,
        "material": material,
        "use_cases": use_cases,
        "packaging": packaging,
        "product_grade_anchor": grade_anchor,
        "comm_imagery_digest": comm_digest,
        "_diagnostics": {
            "raw_rows_count": len(rows),
            "dim_hits_count": len(dim_hits),
            "use_cases_count": len(use_cases),
            "comm_digest_images": comm_digest.get("image_count", 0) if isinstance(comm_digest, dict) else 0,
            "comm_digest_selling_points": len(comm_digest.get("selling_point_phrases", [])) if isinstance(comm_digest, dict) else 0,
        },
    }


def render_summary(sku_truth: Dict[str, Any]) -> str:
    """One-line-per-fact summary suitable for prompt injection."""
    ident = sku_truth.get("identity", {})
    dims = sku_truth.get("dimensions", {})
    mat = sku_truth.get("material", {})
    uc = sku_truth.get("use_cases", [])
    parts = [
        f"Product: {ident.get('product_name_ru', '?')} (archetype={ident.get('archetype', '?')})",
        f"Title: {ident.get('title_ru_full', '?')}",
    ]
    if dims.get("raw_hits"):
        parts.append("Dimensions (verified): " + " / ".join(d["display_ru"] for d in dims["raw_hits"]))
    if mat.get("primary_ru"):
        parts.append(f"Material: {mat['primary_ru']} ({mat.get('finish') or 'finish unspecified'})")
    if uc:
        parts.append("Use cases (from listing): " + " / ".join(u["case_ru"] for u in uc))
    pkg = sku_truth.get("packaging", {})
    if pkg.get("has_real_reference_image"):
        parts.append(f"Packaging refs: {len(pkg['reference_files'])} file(s)")
    else:
        parts.append("Packaging: NO real reference (skip packaging/unboxing slots)")
    # Comm imagery digest
    digest = sku_truth.get("comm_imagery_digest", {})
    if isinstance(digest, dict) and digest.get("image_count"):
        sps = digest.get("selling_point_phrases", []) or []
        if sps:
            parts.append("Verified comm selling points: " + " / ".join(f'"{s}"' for s in sps[:8]))
        sl = digest.get("slot_layout_refs", {}) or {}
        if sl:
            parts.append("Comm imagery has design refs for slots: " + ", ".join(f"{k}({len(v)})" for k, v in sl.items()))
    return "\n".join(parts)


def main() -> None:
    p = argparse.ArgumentParser(description="V7 SKU truth loader — listing.xlsx → sku_truth.yaml")
    p.add_argument("--listing", required=True, help="Path to listing.xlsx (or listing_.xlsx)")
    p.add_argument("--comm-dir", help="Path to 沟通图片 folder (for packaging detection)")
    p.add_argument("--sku-id", help="SKU id (defaults to listing parent dir name)")
    p.add_argument("--out", help="Write YAML to this path (else print JSON to stdout)")
    p.add_argument("--summary", action="store_true", help="Print only the prompt-injection summary")
    args = p.parse_args()

    sku_truth = build_sku_truth(
        Path(args.listing).expanduser().resolve(),
        sku_id=args.sku_id,
        comm_dir=Path(args.comm_dir).expanduser().resolve() if args.comm_dir else None,
    )

    if args.summary:
        print(render_summary(sku_truth))
        return

    if args.out:
        out_path = Path(args.out).expanduser()
        if out_path.suffix in {".yaml", ".yml"}:
            try:
                import yaml  # type: ignore
                out_path.write_text(yaml.safe_dump(sku_truth, allow_unicode=True, sort_keys=False), encoding="utf-8")
            except ImportError:
                out_path.write_text(json.dumps(sku_truth, ensure_ascii=False, indent=2), encoding="utf-8")
        else:
            out_path.write_text(json.dumps(sku_truth, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"wrote {out_path}", file=sys.stderr)
    else:
        print(json.dumps(sku_truth, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
