#!/usr/bin/env python3
"""丝绸生活 沟通图片 文件夹解析器 — xlsx/xls 宽容解析 + 参考图分桶。

输入：类目根目录（含 沟通图片/ 子目录）
输出：dict {
    category: str,                     # 类目名（从路径推断）
    sheet_data: {                      # xlsx 解析结果（可能 None）
        title_ru: str,
        benefits_ru: list[str],        # 5 个卖点（实际 listing 模板就是 5 个，不是 12）
        description_ru: str,
        search_terms_ru: str,
        title_zh: str,
        benefits_zh: list[str],
        description_zh: str,
        competitor_urls: list[str],
    },
    refs: {                            # 参考图分桶
        body: list[Path],              # 主体白底图（命名前缀：主_/main/Main_）
        scene: list[Path],             # 场景图（横屏 / 长宽比 != 1:1）
        poster: list[Path],            # 信息海报 / 详情图（命名前缀：Description_/image_）
    },
    issues: list[str],                 # 警告（缺 xlsx / 解析失败 / 缺主体图 等）
}
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

REF_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
XLSX_EXTS = {".xlsx", ".xls"}
COMM_DIRS = ["沟通图片"]
MAIN_PREFIXES = ["主图_", "主_", "main", "Main_", "main"]


def scan_category(category_dir: Path) -> dict:
    """扫描一个类目根目录，返回 {comm_dir, xlsx, refs[]}."""
    comm_dir = None
    for name in COMM_DIRS:
        cand = category_dir / name
        if cand.exists() and cand.is_dir():
            comm_dir = cand
            break
    if not comm_dir:
        raise SystemExit(f"❌ 未找到沟通图片文件夹: {category_dir}/沟通图片")

    files = [p for p in comm_dir.iterdir() if p.is_file() and not p.name.startswith(".")]
    xlsxs = [p for p in files if p.suffix.lower() in XLSX_EXTS]
    refs = [p for p in files if p.suffix.lower() in REF_EXTS]
    return {
        "category": category_dir.name,
        "comm_dir": comm_dir,
        "xlsx_path": xlsxs[0] if xlsxs else None,
        "ref_paths": refs,
    }


def parse_workbook(xlsx_path: Path) -> dict:
    """宽容解析 xlsx/xls。

    标准 listing 模板：
      行 0: 表头（'属性/Характеристика | 卖点/Преимущества товара | D' 或 'A | B | D'）
      行 1: Заголовок（标题）
      行 2-6: Преимущество 1-5（5 个卖点）
      行 7: Описание（描述）
      行 8: search terms（俄语关键词）
      行 9: JSON（结构化富文本，跳过）
      行 10: 竞品 URL 表头
      行 11: 竞品 URL 实际值
      列 1: 俄语正文
      列 2: 中文翻译

    内容启发式（不依赖列名）：
      - 列 0 含 'Заголовок' → 该行是标题
      - 列 0 含 'Преимущество' → 该行是卖点
      - 列 0 含 'Описание' → 该行是描述
      - 列 0 含 'search terms' → 该行是关键词
      - 列 1 含 'https://www.ozon.ru' → 该行是竞品 URL
    """
    suffix = xlsx_path.suffix.lower()
    rows = []
    if suffix == ".xlsx":
        try:
            import openpyxl  # type: ignore
        except ImportError:
            raise SystemExit("❌ 需要 openpyxl: pip3 install --user openpyxl")
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    elif suffix == ".xls":
        try:
            import xlrd  # type: ignore
        except ImportError:
            raise SystemExit("❌ 需要 xlrd 1.x: pip3 install --user 'xlrd==1.2.0'")
        wb = xlrd.open_workbook(str(xlsx_path))
        ws = wb.sheet_by_index(0)
        rows = [tuple(ws.cell_value(r, c) for c in range(ws.ncols)) for r in range(ws.nrows)]
    else:
        raise ValueError(f"不支持的格式: {suffix}")

    out: dict = {
        "title_ru": "",
        "benefits_ru": [],
        "description_ru": "",
        "search_terms_ru": "",
        "title_zh": "",
        "benefits_zh": [],
        "description_zh": "",
        "competitor_urls": [],
    }

    for row in rows:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = [(str(c).strip() if c is not None else "") for c in row]
        label = cells[0]
        col_b = cells[1] if len(cells) > 1 else ""
        col_c = cells[2] if len(cells) > 2 else ""

        # 内容启发式（标签命中 OR 内容命中）
        if "Заголовок" in label or (label.lower().startswith("title")):
            out["title_ru"] = col_b
            out["title_zh"] = col_c
        elif "Преимущество" in label or label.lower().startswith("преимущество"):
            if col_b:
                out["benefits_ru"].append(col_b)
            if col_c:
                out["benefits_zh"].append(col_c)
        elif "Описание" in label or label.lower().startswith("description"):
            out["description_ru"] = col_b
            out["description_zh"] = col_c
        elif "search terms" in label.lower() or "ключевые" in label.lower():
            out["search_terms_ru"] = col_b
        elif col_b.startswith("http") and "ozon.ru" in col_b:
            # 整行可能都是 URL（竞品行）
            for cell in cells:
                if cell.startswith("http") and "ozon.ru" in cell:
                    out["competitor_urls"].append(cell)

    return out


def classify_refs(ref_paths: list[Path]) -> dict:
    """参考图三通道分桶：body / scene / poster。

    启发式：
    - 文件名前缀 主_/主图_/main/Main_ → body
    - 文件名前缀 Description_/Desc_ → poster
    - 文件名含 image_ / img_ / 数字开头 → 看长宽比+亮度
        - 长宽比 ≈ 1:1（0.85-1.15）+ 平均亮度 > 200 → body
        - 横屏（< 0.7）→ scene
        - 其他 → poster
    """
    try:
        from PIL import Image  # type: ignore
    except ImportError:
        raise SystemExit("❌ 需要 Pillow: pip3 install --user Pillow")

    body: list[Path] = []
    scene: list[Path] = []
    poster: list[Path] = []

    for p in ref_paths:
        name_lower = p.name.lower()

        # 命名前缀启发式
        if any(p.name.startswith(pfx) for pfx in ["主_", "主图_", "main", "Main_", "main_"]):
            body.append(p)
            continue
        if name_lower.startswith(("description_", "desc_")):
            poster.append(p)
            continue

        # 长宽比 + 亮度启发式
        try:
            img = Image.open(p)
            w, h = img.size
            ratio = w / h
            # 缩到 small 算平均亮度（避免大图慢）
            small = img.resize((50, 50)).convert("L")
            avg = sum(small.getdata()) / 2500
        except Exception:
            poster.append(p)
            continue

        if 0.85 <= ratio <= 1.15 and avg > 200:
            body.append(p)
        elif ratio < 0.75:
            scene.append(p)
        else:
            poster.append(p)

    return {"body": body, "scene": scene, "poster": poster}


def parse(category_dir: Path) -> dict:
    """端到端：扫描 → 解析 → 分桶 → 聚合。"""
    scan = scan_category(category_dir)
    issues: list[str] = []

    sheet_data = None
    if scan["xlsx_path"]:
        try:
            sheet_data = parse_workbook(scan["xlsx_path"])
        except Exception as e:
            issues.append(f"xlsx 解析失败 ({scan['xlsx_path'].name}): {e}")
    else:
        issues.append("⚠️  无 xlsx 文件，listing 字段全部缺失，需要人工补")

    refs = classify_refs(scan["ref_paths"])
    if not refs["body"]:
        issues.append("⚠️  缺主体白底图，会用 scene/poster 兜底但产品一致性受影响")

    return {
        "category": scan["category"],
        "comm_dir": str(scan["comm_dir"]),
        "xlsx_path": str(scan["xlsx_path"]) if scan["xlsx_path"] else None,
        "sheet_data": sheet_data,
        "refs": {k: [str(p) for p in v] for k, v in refs.items()},
        "issues": issues,
    }


def main() -> None:
    p = argparse.ArgumentParser(description="解析丝绸生活 沟通图片 文件夹")
    p.add_argument("category_dir", help="类目根目录，例如 /Users/mac/Documents/ozns/丝绸生活/冰箱除味剂")
    p.add_argument("--out", default=None, help="输出 json 路径（默认打印到 stdout）")
    args = p.parse_args()

    cat_dir = Path(args.category_dir).expanduser().resolve()
    if not cat_dir.exists():
        raise SystemExit(f"❌ 类目目录不存在: {cat_dir}")

    result = parse(cat_dir)
    js = json.dumps(result, ensure_ascii=False, indent=2)

    if args.out:
        Path(args.out).write_text(js, encoding="utf-8")
        print(f"✅ {cat_dir.name}: title={result['sheet_data']['title_ru'][:50] if result['sheet_data'] else 'N/A'}... | refs body={len(result['refs']['body'])} scene={len(result['refs']['scene'])} poster={len(result['refs']['poster'])} | {len(result['issues'])} issues")
        for i in result["issues"]:
            print(f"   {i}")
    else:
        print(js)


if __name__ == "__main__":
    main()
